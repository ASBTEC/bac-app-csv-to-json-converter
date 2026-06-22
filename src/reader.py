import openpyxl
from typing import Iterator


def load_workbook(path: str) -> openpyxl.Workbook:
    return openpyxl.load_workbook(path, data_only=True)


def iter_rows(ws: openpyxl.worksheet.worksheet.Worksheet, id_col: str) -> Iterator[dict]:
    headers = [cell.value for cell in ws[1]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = dict(zip(headers, row))
        if not row_dict.get(id_col):
            continue
        yield row_dict
